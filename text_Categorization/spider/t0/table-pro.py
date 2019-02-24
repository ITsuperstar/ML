#coding=utf-8
import os
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import load_workbook
import openpyxl

filename = r'sh.xls'
workbook = xlrd.open_workbook(filename)
sheet = workbook.sheet_by_index(0)
rowNum = sheet.nrows
'''
filename1 = r'sh.xlsx'
workbook1 = xlrd.open_workbook(filename1)
sheet1 = workbook1.sheet_by_index(0)
rowNum1 = sheet1.nrows
newbook1 = copy(workbook1)
newsheet1 = newbook1.get_sheet(0)
'''
#写入已存在的xlsx文件第二种方法
wb = load_workbook("sh.xlsx")#生成一个已存在的wookbook对象
wb1 = wb.active #激活sheet
r=wb1.max_row
for i in range(rowNum):
    wb1.cell(r+i,1,sheet.cell(i,0).value)

print(wb1.max_row)
wb.save("sh.xlsx") #保存

'''
for i in range(rowNum):
    newsheet1.write(rowNum1 + i, 0, sheet.cell(i,0).value)
# 覆盖保存
newbook1.save(filename1)
'''