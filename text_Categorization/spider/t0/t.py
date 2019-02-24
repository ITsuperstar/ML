# encoding=utf-8
import datetime

import jieba

from openpyxl import load_workbook
import openpyxl

#日期的处理
today = datetime.date.today()
date = today + datetime.timedelta(days = -180)
print(date.year)
print(str(date)[5:7]+str(date)[8:10])


detestr = '2018-11-28'
date1 = datetime.datetime.strptime(detestr,'%Y-%m-%d')
date1 = date1 + datetime.timedelta(days = -1)
print(date1)
if date1>(date1 + datetime.timedelta(days = -1)):
    print(date1,'>',(date1 + datetime.timedelta(days = -1)))


#结巴分词
seg_list = jieba.cut("我来到北京清华大学", cut_all=True)
print("Full Mode: " + "/ ".join(seg_list))  # 全模式

seg_list = jieba.cut("我来到北京清华大学", cut_all=False)
print("Default Mode: " + "/ ".join(seg_list))  # 精确模式

seg_list = jieba.cut("他来到了网易杭研大厦")  # 默认是精确模式
print(", ".join(seg_list))

seg_list = jieba.cut_for_search("小明硕士毕业于中国科学院计算所，后在日本京都大学深造")  # 搜索引擎模式
print(", ".join(seg_list))

#表格的处理
# 写入已存在的xlsx文件第一种方法
# class Write_excel(object):
#     '''修改excel数据'''
#     def __init__(self, filename):
#         self.filename = filename
#         self.wb = load_workbook(self.filename)
#         self.ws = self.wb.active  # 激活sheet
#
#     def write(self, row_n, col_n, value):
#         '''写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"'''
#         self.ws.cell(row_n, col_n,value )
#         self.wb.save(self.filename)
#
# we = Write_excel("mylogintest.xlsx")
# we.write(2,2,'pass3')

#写入已存在的xlsx文件第二种方法
wb = load_workbook("sh1.xlsx")#生成一个已存在的wookbook对象
wb1 = wb.active#激活sheet
print(wb1.max_row)
wb1.cell(3,2,'pass2')#往sheet中的第二行第二列写入‘pass2’的数据
print(wb1.cell(3,2).value)
wb.save("sh1.xlsx")#保存