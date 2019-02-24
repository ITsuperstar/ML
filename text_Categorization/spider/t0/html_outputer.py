#coding=utf-8
import os
from openpyxl import load_workbook
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


class HtmlOutputer(object):
    def __init__(self):
        self.datas = []

    def collect_data(self, data):
        if data is None:
            return
        self.datas.append(data)
        if len(self.datas)==100:
            self.excelwrite()
            self.datas = []


    def excelwrite(self):
        filename="sh.xlsx"
        wb = load_workbook(filename)  # 生成一个已存在的wookbook对象
        wb1 = wb.active  # 激活sheet
        r=wb1.max_row+1
        for i in range(len(self.datas)):
            #content = ILLEGAL_CHARACTERS_RE.sub(r'', self.datas[i]['summary'])
            wb1.cell(r + i, 1, self.datas[i]['summary'])  # 往sheet中的第r+i行第1列写入的数据
        print(wb1.max_row)
        wb.save(filename)  # 保存


